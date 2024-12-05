import tkinter as tk
from tkinter import ttk, PhotoImage
from PIL import Image, ImageTk
import requests
import os
import sys
from bs4 import BeautifulSoup
import openpyxl

# Constants
FILE_NAME = "scraped_prices.xlsx"
POKEMON_RED = "#FF0000"
POKEMON_DARK_RED = "#CC0000"
POKEMON_BLUE = "#3B4CCA"
POKEMON_YELLOW = "#FFDE00"
POKEMON_GOLD = "#B3A125"

#Get absolute path to resource, works for PyInstaller
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        # When running as a bundled executable
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# Fetch price from URL
def fetch_grades(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
    except requests.RequestException as e:
        raise ValueError(f"Failed to fetch URL: {e}")

    soup = BeautifulSoup(response.text, 'html.parser')

    grade_ids = ["used_price", "complete_price", "new_price", "graded_price", "box_only_price", "manual_only_price"]
    grades = {}

    # Extract values for each grade
    for grade_id in grade_ids:
        element = soup.find(id=grade_id)
        if element:
            span = element.find("span", class_="price")  # Replace "desired-class-name" with the actual class name
            if span:
                grades[grade_id] = span.text.strip().lstrip('$').replace(',', '')  # Extract and clean the text inside the span
            else:
                grades[grade_id] = None  # If no span is found, store None
        else:
            grades[grade_id] = None  # If not found, store as None

    # Extract item name
    item_name_element = soup.find(id="product_name")
    if item_name_element:
        item_name = ''.join(item_name_element.find_all(string=True, recursive=False)).strip()
    else:
        item_name = "Unknown Item"
        
    return item_name, grades

# Open or create Excel file
def open_or_create_excel():
    try:
        workbook = openpyxl.load_workbook(FILE_NAME)
        if "Scraped Data" in workbook.sheetnames:
            sheet = workbook["Scraped Data"]  # Access the specific sheet
        else:
            # If the sheet doesn't exist, create it
            sheet = workbook.create_sheet("Scraped Data")
            sheet['A1'] = "Item"
            sheet['B1'] = "Ungraded"
            sheet['C1'] = "Grade 7"
            sheet['D1'] = "Grade 8"
            sheet['E1'] = "Grade 9"
            sheet['F1'] = "Grade 9.5"
            sheet['G1'] = "Grade 10"
            sheet['H1'] = "URL"
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Scraped Data"
        sheet['A1'] = "Item"
        sheet['B1'] = "Ungraded"
        sheet['C1'] = "Grade 7"
        sheet['D1'] = "Grade 8"
        sheet['E1'] = "Grade 9"
        sheet['F1'] = "Grade 9.5"
        sheet['G1'] = "Grade 10"
        sheet['H1'] = "URL"
    return workbook, sheet

#check if url exists - needs an open excel file
def check_duplicates(url, sheet):
    # Iterate through the URLs in the last column
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8, values_only=True):
        if row[0] == url:  # If the URL matches
            return True
    return False

# Add new URL data to Excel
def add_new_url(url):
    if not url:
        display_message("Error: Please enter a URL.", "error")
        return

    workbook, sheet = open_or_create_excel()

    # Check for duplicates
    if check_duplicates(url, sheet):
        display_message("Error: URL already exists in the database.", "error")
        root.update_idletasks()  # Ensure the message is rendered immediately
        return

    try:
        item_name, grades = fetch_grades(url)
    except ValueError as e:
        display_message(str(e), "error")
        return

    # Check for grade keys in the grades dictionary and use default `None` if not present
    ungraded = grades.get("used_price", None)
    grade_7 = grades.get("complete_price", None)
    grade_8 = grades.get("new_price", None)
    grade_9 = grades.get("graded_price", None)
    grade_9_5 = grades.get("box_only_price", None)
    grade_10 = grades.get("manual_only_price", None)

    # Append all data as a new row to the Excel sheet
    sheet.append([item_name, ungraded, grade_7, grade_8, grade_9, grade_9_5, grade_10, url])
    save_excel(workbook)
    refresh_table(sheet)
    display_message("New URL added successfully!", "success")

def update_all_prices():
    # Open the Excel file
    workbook, sheet = open_or_create_excel()

    # Get the total number of rows to process
    total_rows = sheet.max_row - 1  # Exclude header row
    if total_rows <= 0:
        display_message("No data to update.", "error")
        return
    
    # Display the progress bar in the instructions area
    progress_bar = ttk.Progressbar(info_frame, length=300, mode="determinate", maximum=total_rows)
    progress_bar.pack(pady=10)

    # Loop through each row starting from the second row (skipping headers)
    for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False), start=1):
        url_cell = row[-1]  # Assume the last column contains the URL
        if url_cell.value:  # Check if the URL cell is not empty
            try:
                # Fetch item name and grades for the URL
                item_name, grades = fetch_grades(url_cell.value)

                # Update each grade column in the corresponding row
                row[0].value = item_name
                row[1].value = grades.get("used_price", None)       # Ungraded
                row[2].value = grades.get("complete_price", None)   # Grade 7
                row[3].value = grades.get("new_price", None)        # Grade 8
                row[4].value = grades.get("graded_price", None)     # Grade 9
                row[5].value = grades.get("box_only_price", None)   # Grade 9.5
                row[6].value = grades.get("manual_only_price", None)  # Grade 10
            except ValueError as e:
                # Skip rows with invalid URLs and log the error
                print(f"Error updating {url_cell.value}: {e}")
                continue

        # Update the progress bar
        progress_bar["value"] = index
        progress_text = f"Updating prices...Processed {index}/{total_rows} rows."
        instructions_label.config(text=progress_text)
        info_frame.update_idletasks()  # Refresh the UI

    # Save the updated Excel file
    save_excel(workbook)

    # Refresh the table display
    refresh_table(sheet)

    # Display success message
    progress_bar.pack_forget()
    display_message("All prices updated successfully!", "success")

# Save Excel file
def save_excel(workbook):
    try:
        workbook.save(FILE_NAME)
    except PermissionError:
        display_message("Error: Permission denied. Close the file and try again.", "error")

# Refresh the table display
def refresh_table(sheet):
    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)

# Display messages in the app
def display_message(message, message_type):
    instructions_label.config(text=message)
    if(message_type=="success"):
        image_label.config(image=squirtle_image)
    else:
        image_label.config(image=pikachu_image)
    root.after(4000, clear_message)

def clear_message():
    instructions_text = """Welcome to the Pokédex Price Scraper!
- Enter a valid PriceCharting URL and select "add URL"
- Use the "Update All Prices" button to update all existing entries
- All data is saved to an excel file named "scraped prices.xlsx" on the "Scraped Data" sheet
"""
    instructions_label.config(text=instructions_text) 
    image_label.config(image=bulbasaur_image) 

# Sorting functionality
def sort_column(tree, col, reverse):
    data = [(tree.set(k, col), k) for k in tree.get_children("")]
    data.sort(reverse=reverse)

    for index, (val, k) in enumerate(data):
        tree.move(k, "", index)

    tree.heading(col, command=lambda: sort_column(tree, col, not reverse))


# Pokédex Colors
POKEDEX_RED = "#FF1C1C"
POKEDEX_BLACK = "#202020"
POKEDEX_GRAY = "#606060"
POKEDEX_YELLOW = "#FFD700"
POKEDEX_BLUE = "#3B4CCA"
POKEDEX_WHITE = "#FFFFFF"

# Set up the GUI
root = tk.Tk()
root.title("Pokédex Price Scraper")
root.geometry("900x700")
root.configure(bg=POKEDEX_RED)

# Modern style
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview", background=POKEDEX_BLUE, fieldbackground=POKEDEX_BLUE, foreground=POKEDEX_WHITE, rowheight=25, font=("Courier New", 10))
style.configure("Treeview.Heading", background=POKEDEX_GRAY, foreground=POKEDEX_YELLOW, font=("Courier New", 10, "bold"))
style.map("Treeview.Heading", background=[("active", POKEDEX_YELLOW)])
style.configure("TButton", background=POKEDEX_YELLOW, foreground=POKEDEX_BLACK, font=("Courier New", 10, "bold"), padding=8)
style.map("TButton", background=[("active", POKEDEX_WHITE)], foreground=[("active", POKEDEX_BLACK)])

# Image and Instructions Frame
info_frame = tk.Frame(root, bg=POKEDEX_GRAY, bd=5, relief="ridge", width=800, height=250)
info_frame.pack(pady=10)  # Use pack for consistent parent layout
info_frame.pack_propagate(False)  # Prevent the frame from resizing to fit its content

# Pikachu Image
image_frame = tk.Frame(info_frame, bg=POKEDEX_RED, width=200, height=200)
image_frame.pack(side="left", padx=10)  # Use pack with side alignment
image_frame.pack_propagate(False)  # Prevent resizing

# Load and resize the image using Pillow
original_image = Image.open(resource_path("happy_bulbasaur.png"))  # Adjust file name as needed
resized_image = original_image.resize((200, 200), Image.LANCZOS)  # Resize to 200x200 pixels
bulbasaur_image = ImageTk.PhotoImage(resized_image)

# Repeat for other images
original_image = Image.open(resource_path("happy_squirtle.png"))
resized_image = original_image.resize((200, 200), Image.LANCZOS)
squirtle_image = ImageTk.PhotoImage(resized_image)

original_image = Image.open(resource_path("surprised_pikachu.png"))
resized_image = original_image.resize((200, 200), Image.LANCZOS)
pikachu_image = ImageTk.PhotoImage(resized_image)

image_label = tk.Label(image_frame, image=bulbasaur_image, bg=POKEDEX_GRAY)
image_label.pack()

# Instructions Label
instructions_text = """Welcome to the Pokédex Price Scraper!
- Enter a valid PriceCharting URL and select "add URL"
- Use the "Update All Prices" button to update all existing entries
- All data is saved to an excel file named "scraped prices.xlsx" on the "Scraped Data" sheet
"""
instructions_label = tk.Label(
    info_frame,
    text=instructions_text,
    bg=POKEDEX_GRAY,
    fg=POKEDEX_WHITE,
    font=("Courier New", 12),
    justify="left",
    wraplength=500
)
instructions_label.pack(side="left", padx=10)  # Use pack with side alignment

# URL input and buttons
input_frame = tk.Frame(root, bg=POKEDEX_BLACK, bd=5, relief="ridge", width=900)
input_frame.pack(pady=15)

tk.Label(input_frame, text="Enter URL:", bg=POKEDEX_BLACK, fg=POKEDEX_WHITE, font=("Courier New", 10)).grid(row=0, column=0, padx=10)
url_entry = tk.Entry(input_frame, width=50, font=("Courier New", 10), bg=POKEDEX_GRAY, fg=POKEDEX_WHITE)
url_entry.grid(row=0, column=1, padx=10, pady=20)

add_button = ttk.Button(input_frame, text="Add New", command=lambda: add_new_url(url_entry.get().strip()))
add_button.grid(row=0, column=2, padx=10)

update_button = ttk.Button(input_frame, text="Update All Prices", command=update_all_prices)
update_button.grid(row=0, column=3, padx=5)

# Table for data display with scrolling
table_frame = tk.Frame(root, bg=POKEDEX_BLACK, bd=5, relief="ridge")
table_frame.pack(pady=10, fill=tk.BOTH, expand=True)

table_scroll_y = tk.Scrollbar(table_frame, orient=tk.VERTICAL)
table_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

table_scroll_x = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
table_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

table = ttk.Treeview(
    table_frame,
    yscrollcommand=table_scroll_y.set,
    xscrollcommand=table_scroll_x.set,
    columns=(
        "item",
        "ungraded_price",
        "grade_seven_price",
        "grade_eight_price",
        "grade_nine_price",
        "grade_nine_half_price",
        "grade_ten_price",
        "url",
    ),
    show="headings",
)
table.heading("item", text="Item", command=lambda: sort_column(table, "item", False))
table.heading("ungraded_price", text="Ungraded", command=lambda: sort_column(table, "ungraded_price", False))
table.heading("grade_seven_price", text="Grade 7", command=lambda: sort_column(table, "grade_seven_price", False))
table.heading("grade_eight_price", text="Grade 8", command=lambda: sort_column(table, "grade_eight_price", False))
table.heading("grade_nine_price", text="Grade 9", command=lambda: sort_column(table, "grade_nine_price", False))
table.heading("grade_nine_half_price", text="Grade 9.5", command=lambda: sort_column(table, "grade_nine_half_price", False))
table.heading("grade_ten_price", text="Grade 10", command=lambda: sort_column(table, "grade_ten_price", False))
table.heading("url", text="URL", command=lambda: sort_column(table, "url", False))

table.column("item", width=200, anchor="w")
table.column("ungraded_price", width=100, anchor="center")
table.column("grade_seven_price", width=100, anchor="center")
table.column("grade_eight_price", width=100, anchor="center")
table.column("grade_nine_price", width=100, anchor="center")
table.column("grade_nine_half_price", width=100, anchor="center")
table.column("grade_ten_price", width=100, anchor="center")
table.column("url", width=300, anchor="w")

table.pack(fill=tk.BOTH, expand=True)

table_scroll_y.config(command=table.yview)
table_scroll_x.config(command=table.xview)

# Initialize table with existing data
workbook, sheet = open_or_create_excel()
refresh_table(sheet)

root.mainloop()

